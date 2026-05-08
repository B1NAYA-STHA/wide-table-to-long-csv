from __future__ import annotations
import re
import sys
import io
from pathlib import Path

import requests
import pandas as pd
from dotenv import load_dotenv
from loguru import logger

sys.path.insert(0, str(Path(__file__).parent))

from parsers.factory import get_layout
from rowllect.warehouse.s3 import upload_df_to_s3

_XLSX_MAGIC = (b"PK\x03\x04", b"PK\x05\x06")


def _fetch(source: str) -> tuple[bytes, str]:
    if source.startswith("http://") or source.startswith("https://"):
        logger.info(f"Downloading {source}")
        resp = requests.get(source, timeout=60)
        resp.raise_for_status()
        filename = source.rstrip("/").split("/")[-1] or "download"
        return resp.content, filename
    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {source}")
    return path.read_bytes(), path.name


def _slug(name: str) -> str:
    stem = Path(name).stem.lower()
    return re.sub(r"[^\w-]", "-", re.sub(r"\s+", "-", stem)).strip("-")


def _sheet_names(content: bytes) -> list[str | None]:
    """Return sheet names for xlsx, or [None] for csv."""
    if content[:4] not in _XLSX_MAGIC:
        return [None]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    return wb.sheetnames


def _upload_to_s3(sheet_slug: str, parsed_df: pd.DataFrame, clean_df: pd.DataFrame, eav_df: pd.DataFrame) -> None:
    """Upload processed CSVs to S3."""
    try:
        key_prefix = f"nhpc/{sheet_slug}"
        upload_df_to_s3(parsed_df, f"{key_prefix}/parsed.csv")
        upload_df_to_s3(clean_df, f"{key_prefix}/clean.csv")
        upload_df_to_s3(eav_df, f"{key_prefix}/eav.csv")
        logger.info(f"[{sheet_slug}] uploaded to S3 → {key_prefix}/")
    except Exception as e:
        logger.error(f"[{sheet_slug}] S3 upload failed: {e}")


def _is_index_sheet(name: str) -> bool:
    n = name.lower()
    return any(w in n for w in ("index", "note", "readme", "cover", "content"))


def _process_sheet(
    content   : bytes,
    sheet_name: str | None,
    prefix    : str,
    out       : Path,
) -> bool:
    try:
        layout   = get_layout(content, sheet_name=sheet_name)
        long_df  = layout.parse(content, sheet_name=sheet_name)
        if long_df.empty:
            logger.warning(f"[{sheet_name}] empty after parse — skipped")
            return False

        clean_df = layout.resolve(long_df)
        if clean_df.empty:
            logger.warning(f"[{sheet_name}] empty after resolve — skipped")
            return False

        eav_df = layout.to_eav(clean_df, indicator_prefix=prefix)

        sheet_out = out / (_slug(sheet_name) if sheet_name else "data")
        sheet_out.mkdir(parents=True, exist_ok=True)
        long_df.to_csv(sheet_out  / "parsed.csv",  index=False, encoding="utf-8-sig")
        clean_df.to_csv(sheet_out / "clean.csv",   index=False, encoding="utf-8-sig")
        eav_df.to_csv(sheet_out   / "eav.csv",     index=False, encoding="utf-8-sig")

        # Upload to S3
        sheet_slug = _slug(sheet_name) if sheet_name else "data"
        _upload_to_s3(sheet_slug, long_df, clean_df, eav_df)

        logger.info(
            f"[{sheet_name or 'sheet'}] {layout.name} → "
            f"parsed={long_df.shape} clean={clean_df.shape} eav={eav_df.shape}"
        )
        return True

    except Exception as e:
        logger.error(f"[{sheet_name}] failed: {e}")
        return False


def run(source: str) -> None:
    content, filename = _fetch(source)
    sheets = _sheet_names(content)
    prefix = f"local/{_slug(filename)}"

    logger.info(f"{filename} — {len(content):,} bytes | {len(sheets)} sheet(s)")

    out = Path("./data") / _slug(filename)
    out.mkdir(parents=True, exist_ok=True)
    (out / f"original{Path(filename).suffix or '.bin'}").write_bytes(content)

    if len(sheets) == 1:
        _process_sheet(content, sheets[0], prefix, out)
    else:
        ok = skipped = failed = 0
        for sheet in sheets:
            if _is_index_sheet(sheet):
                logger.info(f"[{sheet}] skipped (index/notes)")
                skipped += 1
                continue
            if _process_sheet(content, sheet, f"{prefix}/{_slug(sheet)}", out):
                ok += 1
            else:
                failed += 1
        logger.info(f"Done — {ok} succeeded, {skipped} skipped, {failed} failed → {out}/")


if __name__ == "__main__":
    load_dotenv()
    if len(sys.argv) != 2:
        print("Usage: python run_file.py <file_path_or_url>")
        raise SystemExit(1)
    try:
        run(sys.argv[1])
    except Exception as e:
        logger.error(str(e))
        raise SystemExit(1)