from __future__ import annotations
import io
from typing import NamedTuple

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from rowllect.aggregate.location_aggregator import LocationAggregator

from builder.resolve import resolve_districts, COUNTRY
from builder.build_eav import slug, finalize

_COL_PROVINCE  = 0
_COL_DISTRICT  = 1
_COL_PALIKA    = 2
_COL_SEX_START = 3
_VAL_START_1   = _COL_SEX_START + 2  # 1-indexed start of value columns in xlsx

_SEX_LABELS = {"male", "female", "total", "both sexes", "both"}


class _VC(NamedTuple):
    index : int   # 0-indexed column position
    sector: str   # top-level grouping label
    sex   : str   # sex label (empty for no_sex / sex_row variants)


class _Info(NamedTuple):
    sub_layout   : str        # 'sex_paired' | 'sex_row' | 'no_sex'
    breakdown_col: int        # col index carrying row-level breakdown text
    value_cols   : list[_VC]


def detect_value_columns(raw_bytes: bytes) -> _Info:
    """Read xlsx header rows and return column structure info."""
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=False)
    ws = wb.active
    assert ws is not None
    n = ws.max_column

    title_rows = {
        mc.min_row for mc in ws.merged_cells.ranges
        if mc.min_col == 1 and mc.max_col == n
    }
    mmap: dict = {}
    for mc in ws.merged_cells.ranges:
        if mc.min_row in title_rows:
            continue
        val = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                mmap[(r, c)] = val

    def cv(r, c):
        return mmap.get((r, c)) or ws.cell(r, c).value

    def cs(r, c) -> str:
        v = cv(r, c)
        return str(v).strip() if v else ""

    first_data = next(
        (r for r in range(1, ws.max_row + 1)
         if any(isinstance(cv(r, c), (int, float)) for c in range(_VAL_START_1, n + 1))),
        None,
    )
    if first_data is None:
        raise ValueError("No numeric data rows found")

    header_rows = [
        r for r in range(1, first_data)
        if r not in title_rows
        and not any(isinstance(cv(r, c), (int, float)) for c in range(1, n + 1))
        and any(cs(r, c) for c in range(_VAL_START_1, n + 1))
    ]
    if not header_rows:
        raise ValueError("No header rows found")

    row_top = header_rows[-2] if len(header_rows) >= 2 else header_rows[-1]
    row_bot = header_rows[-1]

    bot_labels = {
        cs(row_bot, c).lower()
        for c in range(_VAL_START_1, n + 1)
        if cs(row_bot, c)
    }
    is_sex_paired = bool(bot_labels) and bot_labels.issubset(_SEX_LABELS)

    # Check if col _COL_SEX_START carries sex labels in data rows
    col3_vals: set[str] = set()
    r, sampled = first_data, 0
    while sampled < 30 and r <= ws.max_row:
        v = cs(r, _COL_SEX_START + 1)
        if v and not any(isinstance(cv(r, c), (int, float)) for c in range(_VAL_START_1, n + 1)):
            col3_vals.add(v.lower())
            sampled += 1
        r += 1

    value_cols: list[_VC] = []
    current_group = ""

    if is_sex_paired:
        for c in range(_VAL_START_1, n + 1):
            top, bot = cs(row_top, c), cs(row_bot, c)
            if top:
                current_group = top
            if bot and current_group:
                value_cols.append(_VC(index=c - 1, sector=current_group, sex=bot))
        sub_layout = "sex_paired"
    else:
        for c in range(_VAL_START_1, n + 1):
            top, bot = cs(row_top, c), cs(row_bot, c)
            label = bot or top
            if top and not bot:
                current_group = top
            if label:
                sector = (
                    f"{current_group} - {label}"
                    if current_group and label != current_group
                    else label
                )
                value_cols.append(_VC(index=c - 1, sector=sector, sex=""))
        sub_layout = "sex_row" if bool(col3_vals & _SEX_LABELS) else "no_sex"

    breakdown_col = _COL_SEX_START + 1 if sub_layout == "sex_row" else _COL_SEX_START
    return _Info(sub_layout=sub_layout, breakdown_col=breakdown_col, value_cols=value_cols)


def walk_rows(raw_df: pd.DataFrame, info: _Info) -> pd.DataFrame:
    """Iterate raw xlsx DataFrame and emit one record per (area × value_col)."""
    records  = []
    ctx      = {"province": "", "district": "", "palika": "", "sex": "", "breakdown": ""}
    val_idxs = {vc.index for vc in info.value_cols}
    max_idx  = max(vc.index for vc in info.value_cols)

    for _, row in raw_df.iterrows():
        cells = _cells(row, max_idx + 1)
        if not any(cells):
            continue
        first = str(cells[0] or "").strip()
        if first.lower().startswith("table") or first.lower() == "area":
            continue

        has_values = any(_is_num(cells[i]) for i in val_idxs if i < len(cells))

        if not has_values:
            if cells[_COL_PROVINCE]:
                ctx.update({"province": str(cells[_COL_PROVINCE]).strip(),
                            "district": "", "palika": "", "sex": "", "breakdown": ""})
            elif cells[_COL_DISTRICT]:
                ctx.update({"district": str(cells[_COL_DISTRICT]).strip(),
                            "palika": "", "sex": "", "breakdown": ""})
            elif cells[_COL_PALIKA]:
                ctx.update({"palika": str(cells[_COL_PALIKA]).strip(),
                            "sex": "", "breakdown": ""})
            elif cells[_COL_SEX_START] and info.sub_layout == "sex_row":
                val = str(cells[_COL_SEX_START]).strip()
                if val.lower() in _SEX_LABELS:
                    ctx["sex"]       = val
                    ctx["breakdown"] = ""
                else:
                    ctx["breakdown"] = val
            continue

        breakdown = (
            str(cells[info.breakdown_col]).strip()
            if info.breakdown_col < len(cells) and cells[info.breakdown_col]
            else ctx.get("breakdown", "")
        )

        for vc in info.value_cols:
            if vc.index >= len(cells) or not _is_num(cells[vc.index]):
                continue
            sex = vc.sex if info.sub_layout == "sex_paired" else ctx["sex"]
            records.append({
                "province" : ctx["province"],
                "district" : ctx["district"],
                "palika"   : ctx["palika"],
                "sex"      : sex,
                "breakdown": breakdown,
                "sector"   : vc.sector,
                "value"    : float(str(cells[vc.index]).replace(",", "")),
            })

    return pd.DataFrame(records)


class HierarchicalResolve:

    def resolve(self, long_df: pd.DataFrame) -> pd.DataFrame:
        df = long_df.copy()
        for col in ["province", "district", "palika", "sex", "breakdown", "sector"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
        df["value"] = pd.to_numeric(df["value"], errors="coerce")
        df.dropna(subset=["value"], inplace=True)

        unique_districts = df[df["district"] != ""]["district"].unique()
        dist_to_code     = resolve_districts(unique_districts)
        unmatched        = [n for n in unique_districts if n not in dist_to_code]
        if unmatched:
            logger.warning(f"Unmatched districts (dropped): {unmatched}")

        df["district_code"] = df["district"].map(dist_to_code)
        df["province_code"] = df["district_code"].apply(
            lambda x: x[0] if pd.notna(x) and x else ""
        )
        df["palika_code"] = ""
        for dist_code, grp in df[df["district_code"].notna()].groupby("district_code"):
            palikas = sorted(p for p in grp["palika"].unique() if p)
            pmap    = {p: f"{dist_code}{str(i + 1).zfill(2)}" for i, p in enumerate(palikas)}
            df.loc[grp.index, "palika_code"] = grp["palika"].map(pmap).fillna("")

        return df[~((df["district"] != "") & df["district_code"].isna())].copy()


class HierarchicalEAV:

    def to_eav(self, clean_df: pd.DataFrame, indicator_prefix: str) -> pd.DataFrame:
        combos        = clean_df[["sex", "breakdown", "sector"]].drop_duplicates()
        indicator_map : dict[tuple, str]  = {}
        meta_map      : dict[tuple, list] = {}

        for _, row in combos.iterrows():
            sl, bl, scl = str(row["sex"]).strip(), str(row["breakdown"]).strip(), str(row["sector"]).strip()
            ss, bs, scs = slug(sl), slug(bl), slug(scl)
            key                = (sl, bl, scl)
            indicator_map[key] = "/".join(p for p in [indicator_prefix, ss, bs, scs] if p)
            meta = []
            if sl  and ss : meta.append({"label": sl,  "slug": ss,  "category": "Sex"})
            if bl  and bs : meta.append({"label": bl,  "slug": bs,  "category": "Breakdown"})
            if scl and scs: meta.append({"label": scl, "slug": scs, "category": "Sector"})
            meta_map[key] = meta

        palika_rows = clean_df[clean_df["palika_code"] != ""].copy()
        if palika_rows.empty:
            logger.warning(f"{self.__class__.__name__}: no palika rows found")
            return pd.DataFrame()

        palika_rows["indicator"] = palika_rows.apply(
            lambda r: indicator_map.get(
                (str(r["sex"]).strip(), str(r["breakdown"]).strip(), str(r["sector"]).strip()), ""
            ),
            axis=1,
        )

        eav_input = pd.DataFrame({
            "code"     : palika_rows["palika_code"].values,
            "country"  : COUNTRY,
            "feature"  : "ADM3",
            "indicator": palika_rows["indicator"].values,
            "value"    : palika_rows["value"].values,
        })

        eav_df         = LocationAggregator(eav_input, method="sum", start_level="ADM3").aggregate()
        ind_to_meta    = {v: meta_map[k] for k, v in indicator_map.items()}
        eav_df["Meta"] = eav_df["indicator"].map(ind_to_meta)
        return finalize(eav_df)


def _cells(row: pd.Series, length: int) -> list:
    out = []
    for i in range(length):
        if i >= len(row):
            out.append(None)
            continue
        v = row.iloc[i]
        out.append(
            None if (v is None or (isinstance(v, float) and pd.isna(v)))
            else (str(v).strip() or None)
        )
    return out


def _is_num(val) -> bool:
    if val is None:
        return False
    try:
        float(str(val).replace(",", ""))
        return True
    except ValueError:
        return False