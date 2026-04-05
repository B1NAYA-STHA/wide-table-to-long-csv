"""
rowllect/parsers/hierarchical.py
---------------------------------
Parser for NPHC XLSX tables where Province / District / Palika appear on
their own dedicated area rows (not inline with values).

Two sub-layouts are supported and auto-detected from the column headers:

Sub-layout A — "sex-row" (e.g. indv16: occupation by institutional sector)
---------------------------
  Col 0  Province   — dim row
  Col 1  District   — dim row
  Col 2  Palika     — dim row
  Col 3  Sex        — dim row  (Total / Male / Female)
  Col 4  Breakdown  — inline with values  (occupation category, age group …)
  Col 5+ Values     — one column per sector/indicator  (single header row)

  Row 3: Area | Sex and occupation | Total | Sector …
  Row 4:      |                   |       | Govt | Fin corp | …

  Detected when: the sub-header row (row4) after the value-start column
  contains distinct non-sex labels (sector names, not Male/Female repeats).

Sub-layout B — "sex-paired" (e.g. indv54: literacy by economic activity)
---------------------------
  Col 0  Province   — dim row
  Col 1  District   — dim row
  Col 2  Palika     — dim row
  Col 3  Breakdown  — inline with values  (literacy status, age group …)
  Col 4+ Values     — pairs of (Male, Female) columns per indicator group

  Row 3: Area | Literacy status | Total | Usually Active | Not usually active …
  Row 4:      |                 | Male  | Female | Male  | Female …

  Detected when: the sub-header row (row4) after the value-start column
  consists entirely of Male / Female labels (sex-paired pattern).

Both sub-layouts produce the same long_df schema:
  province | district | palika | sex | breakdown | sector | value

For sub-layout A, sex comes from the dimension row (col 3).
For sub-layout B, sex comes from the column header (Male/Female from row4),
and the sector/indicator name comes from the group label in row3.
"""

from __future__ import annotations

import io
from typing import NamedTuple

import pandas as pd
from openpyxl import load_workbook

from .base import BaseTableParser, TableSchema
from ._detect import read_xlsx_rows, extract_title

# Fixed area column indices — same for every NPHC individual table
_COL_PROVINCE = 0
_COL_DISTRICT = 1
_COL_PALIKA   = 2

# The column where area ends and content begins
_COL_CONTENT_START = 3

_SEX_LABELS = {"male", "female", "total", "both sexes", "both"}


# ---------------------------------------------------------------------------
# Column descriptor
# ---------------------------------------------------------------------------

class ValueColumn(NamedTuple):
    """Describes one value column in the parsed table."""
    index    : int    # 0-based column index in the raw sheet
    sector   : str    # group / indicator name  (from row3)
    sex      : str    # sex label from row4, or '' if layout-A (sex is a dim row)


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

class HierarchicalParser(BaseTableParser):

    def schema(self, raw_bytes: bytes) -> TableSchema:
        title    = extract_title(read_xlsx_rows(raw_bytes)[0])
        col_info = _detect_columns(raw_bytes)
        return TableSchema(
            title       = title,
            subject     = title,
            dim_names   = ["province", "district", "palika", "sex", "breakdown"],
            value_names = list({c.sector for c in col_info.value_cols}),
            layout      = "hierarchical",
            extras      = {
                "col_info"  : col_info,
                "sub_layout": col_info.sub_layout,
            },
        )

    def to_long(self, raw_bytes: bytes, schema: TableSchema) -> pd.DataFrame:
        col_info = schema.extras["col_info"]
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        raw_df = pd.DataFrame(ws.values)
        raw_df = raw_df.dropna(how='all').dropna(axis=1, how='all')  # clean empty rows/cols
        return _parse_rows(raw_df, col_info)


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

class _ColInfo(NamedTuple):
    sub_layout     : str              # 'sex_row' or 'sex_paired'
    breakdown_col  : int              # col index of the breakdown dim
    value_cols     : list[ValueColumn]


def _detect_columns(raw_bytes: bytes) -> _ColInfo:
    """
    Read the header rows and return a _ColInfo describing the table layout.

    Steps:
    1. Load the sheet with merged cells expanded.
    2. Find the two header rows (the last two text-only rows before the first
       numeric data row).
    3. Decide sub-layout from whether the sub-header row is all-sex labels.
    4. Build ValueColumn descriptors for every value column.
    """
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=False)
    ws = wb.active
    n  = ws.max_column

    # Expand merged cells (skip full-row title merges)
    title_rows = {mc.min_row for mc in ws.merged_cells.ranges
                  if mc.min_col == 1 and mc.max_col == n}
    mmap: dict = {}
    for mc in ws.merged_cells.ranges:
        if mc.min_row in title_rows:
            continue
        val = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                mmap[(r, c)] = val

    def cell(r, c):
        return mmap.get((r, c)) or ws.cell(r, c).value

    def cell_str(r, c) -> str:
        v = cell(r, c)
        return str(v).strip() if v else ""

    # Find first data row (first row with a numeric value in value columns)
    first_data = next(
        (r for r in range(1, 40)
         if any(isinstance(cell(r, c), (int, float))
                for c in range(_COL_CONTENT_START + 2, n + 1))),
        None,
    )
    if first_data is None:
        raise ValueError("No numeric data rows found in hierarchical sheet")

    # Collect header rows: text-only rows that have content in the value columns.
    # Rows with content only in the area columns (A-C) are dimension/area rows,
    # not column headers — excluding them is the key to correct sub-layout detection.
    val_start_1idx = _COL_CONTENT_START + 2   # 1-indexed first value column
    header_rows = []
    for r in range(1, first_data):
        if r in title_rows:
            continue
        if any(isinstance(cell(r, c), (int, float)) for c in range(1, n + 1)):
            continue
        # Only count rows that have text in the value-column area
        if any(cell_str(r, c) for c in range(val_start_1idx, n + 1)):
            header_rows.append(r)

    if len(header_rows) < 2:
        raise ValueError(
            f"Expected ≥2 header rows before data row {first_data}, "
            f"found {len(header_rows)}: {header_rows}"
        )

    # The last two header rows define the column names
    row_top = header_rows[-2]   # group-level names  (e.g. "Total", "Usually Active")
    row_bot = header_rows[-1]   # sub-names          (e.g. "Male", "Female" or sector names)

    # Value columns start at _COL_CONTENT_START + 1 (after breakdown dim)
    # Check whether row_bot in the value area is all sex labels → sex-paired
    val_start_col = _COL_CONTENT_START + 2   # 1-indexed for openpyxl
    bot_value_labels = {
        cell_str(row_bot, c).lower()
        for c in range(val_start_col, n + 1)
        if cell_str(row_bot, c)
    }
    is_sex_paired = bool(bot_value_labels) and bot_value_labels.issubset(_SEX_LABELS)

    # Build value columns
    value_cols: list[ValueColumn] = []

    if is_sex_paired:
        # Sub-layout B: row_top = group name (carry-forward), row_bot = Male/Female
        # breakdown is col _COL_CONTENT_START (0-indexed) = col _COL_CONTENT_START+1 (1-indexed)
        current_group = ""
        for c in range(val_start_col, n + 1):   # 1-indexed
            top = cell_str(row_top, c)
            bot = cell_str(row_bot, c)
            if top:
                current_group = top
            if bot and current_group:
                value_cols.append(ValueColumn(
                    index  = c - 1,          # convert to 0-based
                    sector = current_group,
                    sex    = bot,            # 'Male' or 'Female'
                ))
        sub_layout = "sex_paired"

    else:
        # Sub-layout A: row_top = broad group (carry-forward), row_bot = sector names
        # Sex lives in a dimension row (col _COL_CONTENT_START, 0-indexed)
        current_group = ""
        for c in range(val_start_col, n + 1):
            top = cell_str(row_top, c)
            bot = cell_str(row_bot, c)
            label = bot or top
            if top and not bot:
                current_group = top
            if label:
                sector = f"{current_group} - {label}" if current_group and label != current_group else label
                value_cols.append(ValueColumn(
                    index  = c - 1,
                    sector = sector,
                    sex    = "",   # filled from dim row at parse time
                ))
        sub_layout = "sex_row"

    breakdown_col = _COL_CONTENT_START   # 0-based

    return _ColInfo(
        sub_layout    = sub_layout,
        breakdown_col = breakdown_col,
        value_cols    = value_cols,
    )


# ---------------------------------------------------------------------------
# Row walker
# ---------------------------------------------------------------------------

def _parse_rows(raw_df: pd.DataFrame, col_info: _ColInfo) -> pd.DataFrame:
    """
    Walk the data rows maintaining a dimension context and emit one record
    per (province, district, palika, sex, breakdown, sector, value) tuple.

    Works for both sub-layouts:
      sex_row    — sex comes from col _COL_CONTENT_START when no values present
      sex_paired — sex comes from the ValueColumn descriptor
    """
    records = []
    ctx = {"province": "", "district": "", "palika": "", "sex": ""}
    val_indices = {vc.index for vc in col_info.value_cols}
    max_val_idx = max(vc.index for vc in col_info.value_cols)

    for _, row in raw_df.iterrows():
        cells = _row_cells(row, max_val_idx + 1)

        # Skip fully empty rows
        if not any(cells):
            continue

        # Skip title row and column header rows
        first_cell = cells[_COL_PROVINCE] or ""
        if str(first_cell).startswith("Table"):
            continue
        if str(first_cell).strip().lower() == "area":
            continue

        has_values = any(_is_num(cells[i]) for i in val_indices if i < len(cells))

        if not has_values:
            # This is a dimension row — update context
            if cells[_COL_PROVINCE]:
                ctx.update({"province": str(cells[_COL_PROVINCE]).strip(),
                            "district": "", "palika": "", "sex": ""})
            elif cells[_COL_DISTRICT]:
                ctx.update({"district": str(cells[_COL_DISTRICT]).strip(),
                            "palika": "", "sex": ""})
            elif cells[_COL_PALIKA]:
                ctx.update({"palika": str(cells[_COL_PALIKA]).strip(), "sex": ""})
            elif col_info.sub_layout == "sex_row" and cells[col_info.breakdown_col]:
                # In sex_row layout the sex appears in the breakdown col position
                val = str(cells[col_info.breakdown_col]).strip()
                if val.lower() in _SEX_LABELS:
                    ctx["sex"] = val
            continue

        # Value row — emit one record per value column
        breakdown = str(cells[col_info.breakdown_col]).strip() if col_info.breakdown_col < len(cells) else ""

        for vc in col_info.value_cols:
            if vc.index >= len(cells):
                continue
            raw_val = cells[vc.index]
            if not _is_num(raw_val):
                continue

            # Sex: from column descriptor (sex_paired) or from context (sex_row)
            sex = vc.sex if col_info.sub_layout == "sex_paired" else ctx["sex"]

            records.append({
                "province" : ctx["province"],
                "district" : ctx["district"],
                "palika"   : ctx["palika"],
                "sex"      : sex,
                "breakdown": breakdown,
                "sector"   : vc.sector,
                "value"    : float(raw_val),
            })

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _row_cells(row: pd.Series, length: int) -> list:
    """Extract up to `length` cells from a DataFrame row, converting NaN to None."""
    out = []
    for i in range(length):
        if i >= len(row):
            out.append(None)
            continue
        v = row.iloc[i]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            out.append(None)
        else:
            s = str(v).strip()
            out.append(s if s else None)
    return out


def _is_num(val) -> bool:
    if val is None:
        return False
    try:
        float(str(val).replace(",", ""))
        return True
    except ValueError:
        return False