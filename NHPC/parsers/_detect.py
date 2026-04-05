"""
rowllect/parsers/_detect.py
---------------------------
Low-level detection utilities shared by all parsers.
"""

from __future__ import annotations

import io
import re
import csv


# ---------------------------------------------------------------------------
# Cell primitives
# ---------------------------------------------------------------------------

def clean(v) -> str:
    return str(v).strip() if v is not None else ""

def is_numeric(val) -> bool:
    try:
        float(clean(val).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False

def is_all_same(row) -> bool:
    vals = [clean(c) for c in row if clean(c)]
    return len(vals) > 1 and len(set(vals)) == 1

def padded(row, length: int) -> list:
    return list(row) + [None] * max(0, length - len(row))


# ---------------------------------------------------------------------------
# Title extraction
# ---------------------------------------------------------------------------

_TABLE_PREFIX = re.compile(r"^Table\s+\d+[:\.\s]+", re.IGNORECASE)
_SURVEY_TAG   = re.compile(r"[,\s]+(?:NPHC|CBS|Census)\s*\d{4}\s*$", re.IGNORECASE)

def extract_title(rows: list) -> str:
    """
    Return cleaned subject from row 0's first cell.
    Strips 'Table N:' prefix and trailing ', NPHC 2021' tag.
    Returns '' if row 0 has multiple distinct values (it's a header, not a title).
    """
    if not rows or not rows[0]:
        return ""
    first = clean(rows[0][0])
    if not first:
        return ""
    rest = [clean(c) for c in rows[0][1:] if clean(c)]
    if rest:
        return ""
    s = _TABLE_PREFIX.sub("", first)
    s = _SURVEY_TAG.sub("", s)
    return s.strip()


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_csv_rows(path_or_bytes) -> tuple[list, set]:
    """Read CSV from file path or bytes. Returns (rows, title_row_indices)."""
    if isinstance(path_or_bytes, bytes):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = path_or_bytes.decode(enc)
                rows = list(csv.reader(text.splitlines()))
                return rows, set()
            except (UnicodeDecodeError, ValueError):
                continue
        raise ValueError("Cannot decode CSV bytes")

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path_or_bytes, newline="", encoding=enc) as f:
                rows = list(csv.reader(f))
            return rows, set()
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode {path_or_bytes}")


def read_xlsx_rows(path_or_bytes) -> tuple[list, set]:
    """
    Load active XLSX sheet into list-of-lists with merged cells expanded.
    Full-row title merges are left sparse to avoid polluting header detection.
    Returns (rows, title_row_indices_0based).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("pip install openpyxl")

    if isinstance(path_or_bytes, bytes):
        wb = load_workbook(io.BytesIO(path_or_bytes), data_only=True)
    else:
        wb = load_workbook(path_or_bytes, data_only=True)

    ws  = wb.active
    raw = [list(row) for row in ws.iter_rows(values_only=True)]
    n   = ws.max_column

    title_rows = {mc.min_row - 1
                  for mc in ws.merged_cells.ranges
                  if mc.min_col == 1 and mc.max_col == n}

    merge_map: dict = {}
    for mc in ws.merged_cells.ranges:
        if (mc.min_row - 1) in title_rows:
            continue
        val = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row - 1, mc.max_row):
            for c in range(mc.min_col - 1, mc.max_col):
                merge_map[(r, c)] = val

    rows = [
        [merge_map.get((ri, ci), row[ci] if ci < len(row) else None)
         for ci in range(n)]
        for ri, row in enumerate(raw)
    ]
    return rows, title_rows


# ---------------------------------------------------------------------------
# Header block detection
# ---------------------------------------------------------------------------

def detect_header_block(rows: list, title_rows: set,
                        max_scan: int = 15) -> tuple[int, int, int]:
    """
    Return (h_start, h_end, d_start) — all 0-based row indices.
    Header row: ≥2 distinct non-empty cells, not a title/all-same row, <50% numeric.
    Block ends when the next non-empty row is ≥40% numeric.
    """
    def is_header(i, row):
        if i in title_rows or is_all_same(row):
            return False
        vals = [clean(c) for c in row if clean(c)]
        return len(vals) >= 2 and sum(is_numeric(v) for v in vals) / len(vals) < 0.5

    def has_data(row):
        cells = [clean(row[c]) for c in range(1, len(row)) if clean(row[c])]
        return bool(cells) and sum(is_numeric(c) for c in cells) / len(cells) >= 0.4

    candidates = [(i, r) for i, r in enumerate(rows[:max_scan])
                  if any(clean(c) for c in r)]
    h_start = h_end = d_start = None

    for pos, (i, row) in enumerate(candidates):
        if not is_header(i, row):
            continue
        if h_start is None:
            h_start = i
        h_end = i
        if pos + 1 < len(candidates):
            ni, nrow = candidates[pos + 1]
            if not is_header(ni, nrow) and has_data(nrow):
                d_start = ni
                break

    if h_start is None:
        h_start = h_end = (candidates[0][0] if candidates else 0)
    if d_start is None:
        d_start = (h_end or 0) + 1
    return h_start, h_end, d_start


def collapse_headers(rows: list, h_start: int, h_end: int) -> list[str]:
    """
    Combine header rows column-by-column, collapsing adjacent duplicates.
    e.g. rows ['Area','Area','Area'] + ['','',''] -> 'Area' for those cols.
    """
    hrows  = rows[h_start: h_end + 1]
    n_cols = max((len(r) for r in hrows), default=0)
    names  = []
    for c in range(n_cols):
        parts: list[str] = []
        for r in hrows:
            val = clean(r[c]) if c < len(r) else ""
            if val and (not parts or val != parts[-1]):
                parts.append(val)
        names.append(" - ".join(parts))
    return names


# ---------------------------------------------------------------------------
# Column role detection
# ---------------------------------------------------------------------------

_ID_KEYWORDS = re.compile(r"\b(code|_id|\bid\b|no\.|s\.no)\b")
_DESCRIPTIVE = re.compile(
    r"\b(total|male|female|stated|rate|ratio|density|size|growth|"
    r"pop|household|number|count|avg|average|sector|employ)\b"
)

def is_id_col(col_values: list, col_name_str: str = "") -> bool:
    """True when a column holds sequential area/district codes (1..7, 101..709), not measurements."""
    name = col_name_str.lower()
    if _DESCRIPTIVE.search(name):
        return False
    nums = [clean(v) for v in col_values if clean(v) not in ("", "x", "..", "None")]
    if not nums:
        return False
    ints = []
    for v in nums:
        try:
            ints.append(int(v.replace(",", "")))
        except ValueError:
            return False
    if any(i < 0 for i in ints):
        return False
    mn, mx = min(ints), max(ints)
    return bool(_ID_KEYWORDS.search(name)) or (mx <= 1000 and mx - mn <= 900)


def detect_column_roles(col_names: list, data_rows: list,
                        n_sample: int = 40) -> tuple[list, list, list, set]:
    """Return (dim_cols, value_cols, ratios, id_cols)."""
    THRESHOLD = 0.6
    sample = [r for r in data_rows[:n_sample] if any(clean(c) for c in r)]
    n_cols = max(len(col_names), max((len(r) for r in sample), default=0))

    def _ratio(c):
        cells = [clean(r[c]) for r in sample
                 if c < len(r) and clean(r[c]) not in ("", "None")]
        return sum(is_numeric(x) for x in cells) / len(cells) if cells else 0.0

    ratios = [_ratio(c) for c in range(n_cols)]

    id_cols = {
        c for c in range(n_cols)
        if ratios[c] >= THRESHOLD
        and is_id_col([clean(r[c]) for r in sample if c < len(r)],
                      col_names[c] if c < len(col_names) else "")
    }

    value_start = next(
        (c for c in range(n_cols) if ratios[c] >= THRESHOLD and c not in id_cols),
        None,
    )
    if value_start is None:
        raise ValueError(
            f"No numeric value columns found.\n"
            f"Ratios: {[round(r, 2) for r in ratios]}  ID cols: {sorted(id_cols)}"
        )

    dim_cols = list(range(value_start))
    value_cols = [
        c for c in range(value_start, n_cols)
        if c not in id_cols
        and (ratios[c] >= THRESHOLD
             or any(ratios[v] >= THRESHOLD
                    for v in range(c, min(c + 3, n_cols))))
    ]
    return dim_cols, value_cols, ratios, id_cols


def detect_layout(data_rows: list, dim_cols: list, value_cols: list,
                  n_sample: int = 80) -> str:
    """Return 'hierarchical' if any dim column never co-appears with values, else 'flat'."""
    if not (data_rows and dim_cols and value_cols):
        return "flat"
    width  = max(value_cols) + 2
    sample = [padded(r, width) for r in data_rows[:n_sample]
              if any(clean(c) for c in r)]
    for dc in dim_cols:
        filled = [r for r in sample if clean(r[dc]) and not is_numeric(clean(r[dc]))]
        if filled and not any(
            is_numeric(clean(r[c])) for r in filled
            for c in value_cols if c < len(r)
        ):
            return "hierarchical"
    return "flat"