"""
convert.py  —  Wide census table → long (tidy) CSV

Three layouts detected automatically:

  Flat A   One row per area:
             prov_dist_code | prov_dist_name | households | total | male …

  Flat B   2-3 rows per area (Total / Male / Female):
             Area  | Category | Pop 5+ | Can read | …
             Nepal | Total    | 26725  | …

  Hierarchical   Multi-row merged headers; Area / Sex / Category each
                 appear on their own row, not inline with values.

Output:  <dim1> | <dim2> | … | Indicator | Value

Usage:
  python convert.py input.csv          output.csv
  python convert.py input.xlsx         output.csv
  python convert.py inputs/            outputs/   --batch
  python convert.py tricky.csv         out.csv    --value-start-col 3
"""

import csv
import os
import re
import argparse


# ---------------------------------------------------------------------------
# Primitives
# ---------------------------------------------------------------------------

def clean(v):
    return str(v).strip() if v is not None else ""

def is_numeric(val):
    try:
        float(clean(val).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False

def is_all_same(row):
    """True when every non-empty cell in a row is identical (title row)."""
    vals = [clean(c) for c in row if clean(c)]
    return len(vals) > 1 and len(set(vals)) == 1

def col_name(col_names, c, fallback="Not stated"):
    n = col_names[c].strip() if c < len(col_names) else ""
    return n or fallback

def padded(row, length):
    return list(row) + [None] * max(0, length - len(row))


# ---------------------------------------------------------------------------
# ID-column detection
# ---------------------------------------------------------------------------

_ID_KEYWORDS    = re.compile(r"\b(code|_id|\bid\b|no\.|s\.no)\b")
_DESCRIPTIVE    = re.compile(
    r"\b(total|male|female|stated|rate|ratio|density|size|growth|"
    r"pop|household|number|count|avg|average|sector|employ)\b"
)

def is_id_col(col_values, col_name_str=""):
    """
    True when a column is a sequential ID / district code, not a measurement.

    Rules:
      - All non-empty values must be positive integers.
      - Header contains an ID keyword  OR  values form a compact range
        (max ≤ 1000, range ≤ 900) — province codes 1..7, district 101..709.
      - Any descriptive word in the header disqualifies it immediately.
    """
    name = col_name_str.lower()
    if _DESCRIPTIVE.search(name):
        return False

    nums = [clean(v) for v in col_values if clean(v) not in ("", "x", "..", "None")]
    if not nums:
        return False

    ints = []
    for v in nums:
        try:
            ints.append(int(v.replace(",", "")))
        except ValueError:
            return False                        # non-integer → not an ID col

    if any(i < 0 for i in ints):
        return False

    mn, mx = min(ints), max(ints)
    return bool(_ID_KEYWORDS.search(name)) or (mx <= 1000 and mx - mn <= 900)


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------

def read_csv(path):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.reader(f)), set()
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode {path}")


def read_xlsx(path):
    """
    Load the active sheet into a list-of-lists.

    Merged cells are expanded so every cell in the range carries the value,
    EXCEPT for full-row title merges (row spans all columns) — those are
    left sparse to avoid the title text polluting column-name detection.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    raw    = [list(row) for row in ws.iter_rows(values_only=True)]
    n_cols = max((len(r) for r in raw), default=0)

    # Rows whose merge spans every column → title rows; skip their expansion
    title_rows = {mc.min_row - 1
                  for mc in ws.merged_cells.ranges
                  if mc.min_col == 1 and mc.max_col == ws.max_column}

    merge_map = {}
    for mc in ws.merged_cells.ranges:
        if (mc.min_row - 1) in title_rows:
            continue
        val = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row - 1, mc.max_row):
            for c in range(mc.min_col - 1, mc.max_col):
                merge_map[(r, c)] = val

    rows = [
        [merge_map.get((ri, ci), row[ci] if ci < len(row) else None)
         for ci in range(n_cols)]
        for ri, row in enumerate(raw)
    ]
    return rows, title_rows


# ---------------------------------------------------------------------------
# Step 1 — locate the header block
# ---------------------------------------------------------------------------

def detect_header_block(rows, title_rows=None, max_scan=15):
    """
    Return (h_start, h_end, d_start) — all 0-based.

    Header row criteria:
      ≥ 2 distinct non-empty cells, not a title row, < 50 % numeric.
    The block ends when the following non-empty row is mostly numeric.
    """
    title_rows = title_rows or set()

    def is_header(i, row):
        if i in title_rows or is_all_same(row):
            return False
        vals = [clean(c) for c in row if clean(c)]
        return len(vals) >= 2 and sum(is_numeric(v) for v in vals) / len(vals) < 0.5

    def has_data(row):
        cells = [clean(row[c]) for c in range(1, len(row)) if clean(row[c])]
        return bool(cells) and sum(is_numeric(c) for c in cells) / len(cells) >= 0.4

    candidates = [(i, r) for i, r in enumerate(rows[:max_scan]) if any(clean(c) for c in r)]
    h_start = h_end = d_start = None

    for pos, (i, row) in enumerate(candidates):
        if not is_header(i, row):
            continue
        if h_start is None:
            h_start = i
        h_end = i
        if pos + 1 < len(candidates):
            ni, nrow = candidates[pos + 1]
            if not is_header(ni, nrow) and has_data(nrow):
                d_start = ni
                break

    if h_start is None:
        h_start = h_end = (candidates[0][0] if candidates else 0)
    if d_start is None:
        d_start = (h_end or 0) + 1
    return h_start, h_end, d_start


# ---------------------------------------------------------------------------
# Step 2 — flatten multi-row headers into column names
# ---------------------------------------------------------------------------

def collapse_headers(rows, h_start, h_end):
    """
    Combine header rows column-by-column, collapsing adjacent duplicates.

    Example (after merge expansion):
      row 0: ['Area','Area','Area','Sex','Sex','Total','Sector','Sector',…]
      row 1: ['Area','Area','Area','Sex','Sex','Total','Govt','Fin corp',…]
      → col 6 = 'Sector - Govt',  col 5 = 'Total',  col 3 = 'Sex'
    """
    hrows  = rows[h_start: h_end + 1]
    n_cols = max((len(r) for r in hrows), default=0)
    names  = []
    for c in range(n_cols):
        parts = []
        for r in hrows:
            val = clean(r[c]) if c < len(r) else ""
            if val and (not parts or val != parts[-1]):
                parts.append(val)
        names.append(" - ".join(parts))
    return names


# ---------------------------------------------------------------------------
# Step 3 — classify columns as dimension or value
# ---------------------------------------------------------------------------

def detect_column_roles(col_names, data_rows, n_sample=40):
    """
    Return (dim_cols, value_cols, ratios, id_cols).

    Algorithm:
      1. Compute numeric-cell ratio per column on a data sample.
      2. Mark numeric columns that pass is_id_col() as ID columns
         (kept as dimensions, not values).
      3. value_start = leftmost numeric, non-ID column.
      4. dim_cols   = everything left of value_start.
      5. value_cols = numeric (or sparse-but-flanked) columns from
                      value_start onward, excluding ID columns.
    """
    THRESHOLD = 0.6
    sample = [r for r in data_rows[:n_sample] if any(clean(c) for c in r)]
    n_cols = max(len(col_names), max((len(r) for r in sample), default=0))

    ratios = [
        (lambda cells: sum(is_numeric(c) for c in cells) / len(cells) if cells else 0.0)(
            [clean(r[c]) for r in sample if c < len(r) and clean(r[c]) not in ("", "None")]
        )
        for c in range(n_cols)
    ]

    id_cols = {
        c for c in range(n_cols)
        if ratios[c] >= THRESHOLD
        and is_id_col([clean(r[c]) for r in sample if c < len(r)],
                      col_names[c] if c < len(col_names) else "")
    }

    value_start = next(
        (c for c in range(n_cols) if ratios[c] >= THRESHOLD and c not in id_cols),
        None,
    )
    if value_start is None:
        raise ValueError(
            "No numeric value columns found.\n"
            f"Ratios: {[round(r, 2) for r in ratios]}  ID cols: {sorted(id_cols)}\n"
            "Use --value-start-col to set manually."
        )

    dim_cols   = list(range(value_start))
    value_cols = [
        c for c in range(value_start, n_cols)
        if c not in id_cols
        and (ratios[c] >= THRESHOLD
             or any(ratios[v] >= THRESHOLD for v in range(c, min(c + 3, n_cols))))
    ]
    return dim_cols, value_cols, ratios, id_cols


# ---------------------------------------------------------------------------
# Step 4 — detect flat vs hierarchical layout
# ---------------------------------------------------------------------------

def detect_layout(data_rows, dim_cols, value_cols, n_sample=80):
    """
    Return 'hierarchical' if any dim column never appears in the same row
    as value cells — meaning it only populates dedicated dimension rows.
    Otherwise return 'flat'.
    """
    if not (data_rows and dim_cols and value_cols):
        return "flat"

    width  = max(value_cols) + 2
    sample = [padded(r, width) for r in data_rows[:n_sample] if any(clean(c) for c in r)]

    for dc in dim_cols:
        filled = [r for r in sample if clean(r[dc]) and not is_numeric(clean(r[dc]))]
        if filled and not any(
            is_numeric(clean(r[c])) for r in filled for c in value_cols if c < len(r)
        ):
            return "hierarchical"
    return "flat"


# ---------------------------------------------------------------------------
# Melt helpers
# ---------------------------------------------------------------------------

def _indicator_map(col_names, value_cols, unnamed_label):
    return {c: col_name(col_names, c, unnamed_label) for c in value_cols}

def _update_ctx(ctx, dim_cells):
    """Clear context at and below the leftmost changed level, then update."""
    leftmost = min(dim_cells)
    ctx = {k: v for k, v in ctx.items() if k < leftmost}
    ctx.update(dim_cells)
    return ctx


# ---------------------------------------------------------------------------
# Melt — flat layout (A / B)
# ---------------------------------------------------------------------------

def melt_flat(data_rows, col_names, dim_cols, value_cols, unnamed_label="Not stated"):
    """
    Standard wide-to-long pivot.
    Blank dim cells carry forward (handles merged-cell patterns).
    Single-token non-numeric rows (footers) are skipped.
    """
    indic  = _indicator_map(col_names, value_cols, unnamed_label)
    width  = max(value_cols) + 1 if value_cols else 1
    last   = {c: "" for c in dim_cols}
    out    = []

    for row in data_rows:
        p  = padded(row, width)
        ne = [clean(p[c]) for c in range(len(p)) if clean(p[c])]
        if len(ne) == 1 and not is_numeric(ne[0]):   # footer / note row
            continue
        for c in dim_cols:
            if clean(p[c]):
                last[c] = clean(p[c])
        if not any(clean(p[c]) for c in value_cols):
            continue
        dim_vals = [last[c] for c in dim_cols]
        for c in value_cols:
            if clean(p[c]):
                out.append(dim_vals + [indic[c], clean(p[c])])

    dim_headers = [col_name(col_names, c, f"dim_{c}") for c in dim_cols]
    return dim_headers + ["Indicator", "Value"], out


# ---------------------------------------------------------------------------
# Melt — hierarchical layout (C)
# ---------------------------------------------------------------------------

def melt_hierarchical(data_rows, col_names, dim_cols, value_cols,
                      unnamed_label="Not stated"):
    """
    Walks rows tracking a dimension context.

    Pure-dim rows (text in dim cols, no values) → update context,
    resetting deeper levels when a higher-level dim changes.

    Value rows → optionally update context with any inline dim cells,
    then emit one output record per value column.
    """
    indic  = _indicator_map(col_names, value_cols, unnamed_label)
    width  = max(value_cols) + 2 if value_cols else 2
    ctx    = {}
    out    = []

    for row in data_rows:
        p = padded(row, width)
        if not any(clean(c) for c in p):
            continue

        dim_cells = {c: clean(p[c]) for c in dim_cols
                     if clean(p[c]) and not is_numeric(clean(p[c]))}
        val_cells = {c: clean(p[c]) for c in value_cols if clean(p[c])}

        # Lone non-numeric token with no dim match → footer, skip
        ne = [clean(p[c]) for c in range(len(p)) if clean(p[c])]
        if not val_cells and len(ne) == 1 and not dim_cells:
            continue

        if dim_cells:
            ctx = _update_ctx(ctx, dim_cells)
        if val_cells:
            dim_tuple = [ctx.get(c, "") for c in sorted(dim_cols)]
            for c, raw in val_cells.items():
                out.append(dim_tuple + [indic[c], raw])

    # Deduplicate repeated header names with level suffixes
    raw  = [col_name(col_names, c, f"dim_{c}") for c in sorted(dim_cols)]
    freq = {}
    for n in raw:
        freq[n] = freq.get(n, 0) + 1
    seen, hdrs = {}, []
    for n in raw:
        if freq[n] == 1:
            hdrs.append(n)
        else:
            seen[n] = seen.get(n, 0) + 1
            hdrs.append(f"{n} (level {seen[n]})")

    return hdrs + ["Indicator", "Value"], out


# ---------------------------------------------------------------------------
# Top-level convert
# ---------------------------------------------------------------------------

def convert(input_path, output_path,
            dim_cols=None, value_start_col=None, skip_rows=None,
            unnamed_label="Not stated", verbose=True):
    """Convert one file (CSV or XLSX) to long-format CSV."""

    # Read
    ext = os.path.splitext(input_path)[1].lower()
    rows, title_rows = (read_xlsx if ext in (".xlsx", ".xlsm", ".xls") else read_csv)(input_path)
    if not rows:
        raise ValueError(f"Empty file: {input_path}")

    # Header block
    if skip_rows is not None:
        h_start = h_end = skip_rows
        d_start = skip_rows + 1
    else:
        h_start, h_end, d_start = detect_header_block(rows, title_rows)

    col_names = collapse_headers(rows, h_start, h_end)
    data_rows = [r for r in rows[d_start:] if any(clean(c) for c in r)]
    if not data_rows:
        raise ValueError("No data rows found.")

    # Column roles — auto-detect then apply any manual overrides
    _dim, _val, ratios, _id = detect_column_roles(col_names, data_rows)
    if dim_cols is not None:
        _dim = list(dim_cols)
    if value_start_col is not None:
        _val = list(range(value_start_col, len(col_names)))

    layout = detect_layout(data_rows, _dim, _val)

    if verbose:
        print(f"  File     : {os.path.basename(input_path)}")
        print(f"  Headers  : rows {h_start}–{h_end}, data from row {d_start}")
        print(f"  Columns  : {col_names}")
        print(f"  Dim cols : {_dim}  |  ID cols: {sorted(_id)}  |  Value cols: {_val}")
        print(f"  Layout   : {layout}")
        if ratios:
            print(f"  Ratios   : {[round(r, 2) for r in ratios]}")

    # Melt
    if layout == "hierarchical":
        hdr, out_rows = melt_hierarchical(data_rows, col_names, _dim, _val, unnamed_label)
    else:
        hdr, out_rows = melt_flat(data_rows, col_names, _dim, _val, unnamed_label)

    # Write
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(hdr)
        writer.writerows(out_rows)

    if verbose:
        print(f"  Rows     : {len(out_rows)}")
        print(f"  → {output_path}\n")

    return hdr, out_rows


# ---------------------------------------------------------------------------
# Batch mode
# ---------------------------------------------------------------------------

def convert_folder(input_dir, output_dir, **kwargs):
    EXTS  = {".csv", ".xlsx", ".xlsm", ".xls"}
    files = sorted(f for f in os.listdir(input_dir)
                   if os.path.splitext(f)[1].lower() in EXTS)
    if not files:
        print("No files found in", input_dir)
        return

    ok, failed = 0, []
    for fn in files:
        stem = os.path.splitext(fn)[0]
        print(f"\n▶ {fn}")
        try:
            convert(os.path.join(input_dir, fn),
                    os.path.join(output_dir, f"long_{stem}.csv"),
                    **kwargs)
            ok += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            failed.append(fn)

    print(f"\n✓ {ok} converted, ✗ {len(failed)} failed")
    if failed:
        print("  Failed:", failed)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(
        description="Convert wide census CSV/XLSX to long (tidy) format.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
layouts (auto-detected)
  flat           all dims appear in the same row as values;
                 ID/code columns (1..7, 101..709) kept as dims
  hierarchical   dim rows are separate from value rows;
                 multi-row merged headers collapsed automatically

overrides (when auto-detection is wrong)
  --skip-rows N          header starts at row N (0-based)
  --dim-cols 0 1 2       force dimension column indices
  --value-start-col 3    force first value column index

examples
  python convert.py table01.csv        output.csv
  python convert.py indv64-koshi.xlsx  output.csv
  python convert.py inputs/            outputs/   --batch
  python convert.py tricky.csv         out.csv    --value-start-col 3
        """,
    )
    p.add_argument("input",               help="Input file or folder (with --batch)")
    p.add_argument("output",              help="Output CSV or folder (with --batch)")
    p.add_argument("--dim-cols",          type=int, nargs="+")
    p.add_argument("--value-start-col",   type=int)
    p.add_argument("--skip-rows",         type=int)
    p.add_argument("--unnamed-label",     default="Not stated")
    p.add_argument("--batch",             action="store_true")
    p.add_argument("--quiet",             action="store_true")
    args = p.parse_args()

    kwargs = dict(
        dim_cols        = args.dim_cols,
        value_start_col = args.value_start_col,
        skip_rows       = args.skip_rows,
        unnamed_label   = args.unnamed_label,
        verbose         = not args.quiet,
    )
    if args.batch:
        convert_folder(args.input, args.output, **kwargs)
    else:
        convert(args.input, args.output, **kwargs)


if __name__ == "__main__":
    main()